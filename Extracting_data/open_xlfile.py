

from openpyxl import load_workbook

data_file = 'data/mapping_police_violence_snapshot_061920.xlsx'

## Load the entire workbook

wb = load_workbook(data_file)

## List all the sheets in the file
print("Found the following worksheets:")
for sheetname in wb.sheetnames:
	print(sheetname)




